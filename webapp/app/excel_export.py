import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(10, length + 2)


def export_document(title, doc_id, doc_date, employee_label, lines):
    """Build a single document (voucher-style) workbook: title + header fields + line table.

    `lines` is a list of dicts with keys LineNum, ProductId, ProductName, Quantity.
    Returns a BytesIO positioned at 0, ready to send as a file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = doc_id[:31]

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "單號"
    ws["A3"].font = LABEL_FONT
    ws["B3"] = doc_id
    ws["A4"] = "日期"
    ws["A4"].font = LABEL_FONT
    ws["B4"] = str(doc_date)
    ws["A5"] = "經手員工"
    ws["A5"].font = LABEL_FONT
    ws["B5"] = employee_label

    header_row = 7
    columns = ["行號", "物料代號", "物料名稱", "數量"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_idx = header_row + 1
    for line in lines:
        ws.cell(row=row_idx, column=1, value=line["LineNum"])
        ws.cell(row=row_idx, column=2, value=line["ProductId"])
        ws.cell(row=row_idx, column=3, value=line["ProductName"])
        ws.cell(row=row_idx, column=4, value=float(line["Quantity"]))
        row_idx += 1

    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_table(sheet_title, columns, rows):
    """Flat tabular export. `columns` is a list of (header_label, dict_key).
    `rows` is a list of dict-like records.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(columns, start=1):
            value = row[key]
            if hasattr(value, "__float__") and not isinstance(value, (int, float)):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
